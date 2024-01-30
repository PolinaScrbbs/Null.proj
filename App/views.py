from django.shortcuts import render, redirect
from Auth.models import CustomUser, Role
from .models import Direction, Event, Participant
from .forms import EventForm, PhoneNumberForm
from django.http import HttpResponse
from django.http import JsonResponse
from openpyxl import load_workbook

def index(request): 
    if request.user.is_authenticated:
        
        directions = Direction.objects.all()
        events = Event.objects.all() 
 
        context = { 
            'title': 'Главная страница', 
            'directions': directions, 
            'events': events, 
        }

        return render(request, 'index.html', context=context) 
    else: 
        return redirect('registration')

def event(request, title):
    user = request.user
    event = Event.objects.get(title=title)
    try:
        Participant.objects.get(event=event, participant=user)
        is_reg = True   
    except:
        is_reg = False

    context = {
        'title': f'Событие {title}',
        'event': event,
        'is_reg': is_reg
    }

    return render(request, 'event.html', context)

def reg_event(request, event_id):
    user = request.user
    event = Event.objects.get(id=event_id)

    try:
        Participant.objects.get(event=event, participant=user)
        title = Event.objects.get(id=event_id).title
        print(title)
        return redirect('event', title)
    except:
        Participant.objects.create(event=event, participant=user)
        title = Event.objects.get(id=event_id).title
        return redirect('event', title)
    
def event_form(request):
    if request.method == 'POST':
        form = EventForm(request.user, request.POST)
        if form.is_valid():
            # Делаем что-то с формой, например, сохраняем событие
            event = form.save(commit=False)
            event.save()
            return redirect('index')
    else:
        form = EventForm(request.user)

    return render(request, 'event_form.html', {'form': form})

def profile(request, username):
    user = request.user
    role = Role.objects.get(id=user.role_id)
    phone_number_form = PhoneNumberForm
    if request.method == 'POST':
        phone_number_form = PhoneNumberForm(request.POST)
        if phone_number_form.is_valid():
            phone_number = phone_number_form.cleaned_data['phone_number']
            CustomUser.objects.filter(id=user.id).update(phone_number=phone_number)

    # try:
    #     direction = Direction.objects.get(id=user.direction_id)
    # except Direction.DoesNotExist:
    #     direction = None  # или предпочтительный способ обработки отсутствия объекта

    # try:
    #     event = Event.objects.get(id=user.event_id)
    # except Event.DoesNotExist:
    #     event = None  # или предпочтительный способ обработки отсутствия объекта
        
    context={
        'title': 'Профиль',
        'user': user,
        'role': role,
        'phone_number_form': phone_number_form
        # 'direction': direction,
        # 'event': event
    }

    return render(request, 'profile.html', context=context)

def upload_avatar(request):
    if request.method == 'POST' and request.FILES.get('file'):
        user = CustomUser.objects.get(full_name=request.user)
        if user.is_authenticated:
            uploaded_file = request.FILES['file']
            user.photo = uploaded_file
            user.save()

            # Возвращаем JSON с URL новой фотографии
            response_data = {'url': user.photo.url}
            return JsonResponse(response_data)

    # Если что-то пошло не так, возвращаем ошибку
    return JsonResponse({'error': 'Invalid request'})

def import_file_form(request):
    context = {
        'title': 'Импорт'
    }
    
    return render(request, 'import.html', context)

def add_event(sheet,row_num,  start_column):

    # Цикл для чтения данных до окончания данных в первом столбце
    while sheet.cell(row=row_num, column=start_column).value:
        # Создаем объект Event, используя данные из текущей строки
        data = {
            'title': sheet.cell(row=row_num, column=1).value,
            'date': sheet.cell(row=row_num, column=2).value,
            'direction': Direction.objects.get(id = sheet.cell(row=row_num, column=3).value)
        }
                
        Event.objects.create(**data)

        # Переходим к следующей строке
        row_num += 1
                
def add_user(sheet, row_num,  start_column):

    # Цикл для чтения данных до окончания данных в первом столбце
    while sheet.cell(row=row_num, column=start_column).value:
        # Создаем объект Event, используя данные из текущей строки
        data = {
            'full_name': sheet.cell(row=row_num, column=1).value,
            'password': sheet.cell(row=row_num, column=2).value,
            'is_superuser': False,
            'email': sheet.cell(row=row_num, column=3).value,
            'phone_number':  sheet.cell(row=row_num, column=4).value,
            'photo': None,
            'is_active': True,
            'is_staff': True,
            'role_id': 3
        }
                
        CustomUser.objects.create(**data)

        # Переходим к следующей строке
        row_num += 1
        
def import_file(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        model = request.POST.get('dropdown')

        try:
            # Загружаем книгу Excel
            workbook = load_workbook(uploaded_file, data_only=True)

            # Получаем активный лист
            sheet = workbook.active

            # Указываем диапазон столбцов для чтения
            start_column = 1  # с первого столбца

            # Указываем переменные для цикла
            row_num = 1  # начиная с первой строки

            if model == 'CustomUser':
                add_user(sheet, row_num, start_column)
            else:
                add_event(sheet, row_num, start_column)
                
            return redirect('import_file_form')

        except Exception as e:
            print(f"Error processing Excel file: {e}")
            return JsonResponse({'status': 'error', 'message': 'Error processing Excel file'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})


    # def my_people(request):
#         participants = Participant.objects.all()
#         events = Event.objects.all() 
#         context = { 
#             'title': 'Участники',
#             'participants': participants, 
#             'events': events, 
#         }
#         return render(request, 'my_people.html', context=context) 
    
# def my_info(request):
#         user = request.user
#         user_events = Event.objects.filter(direction__creator=user)
#         directions = Direction.objects.all()
#         context = { 
#             'title': 'Мероприятия',
#             'user_events': user_events,
#             'directions': directions
#         }
#         return render(request, 'my_info.html', context=context) 
    
# def jury(request):
#         juries = CustomUser.objects.filter(role_id = 2)
#         events = Event.objects.all()
#         context = { 
#             'title': 'Жюри',
#             'juries': juries,
#             'events': events,
#         }
#         return render(request, 'jury.html', context=context) 